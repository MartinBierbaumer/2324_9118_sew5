import argparse
import logging
import sys
import openpyxl
import unicodedata
import random
import logging.handlers

verbose = False
quiet = False

logger = logging.getLogger(__name__)
handler = logging.handlers.RotatingFileHandler("create_class.log", maxBytes=10000, backupCount=5)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)

stream_handler = logging.StreamHandler(sys.stdout)
stream_handler.setFormatter(formatter)

logger.addHandler(handler)
logger.addHandler(stream_handler)

def replace_normalize(s):
    replacements = {
        'ä': 'ae',
        'ö': 'oe',
        'ü': 'ue',
        'Ä': 'Ae',
        'Ö': 'Oe',
        'Ü': 'Ue',
        'ß': 'ss'
    }
    for umlaut, replacement in replacements.items():
        s = s.replace(umlaut, replacement)
    return unicodedata.normalize('NFC', unicodedata.normalize('NFD', ''.join(char for char in s)))

def create_delete_user(name, create, delete):
    logger.info(f"Creating user: {name}")
    if verbose: print(f"echo creating user: {name}", file=create)

    command = f"""
getent passwd {name} > /dev/null && echo '{name} existiert schon, Abbrcuh' && exit 1
groupadd {name}
useradd -m -d /home/{name} -c {name} -g {name} -s /bin/bash {name}
echo {name}:{name} | chpasswd"""

    print(command, file=create)

    logger.info(f"Lösche {name}")
    if verbose: print(f"lösche {name}", file=delete)
    print(f"userdel -r {name}", file=delete)



def create_user(name, create, pwd):
    logger.info(f"Erstelle {name}")
    if verbose: print(f"echo erstelle {name}", file=create)

    name = replace_normalize(name).lower()
    pwd = pwd.replace("'", "\\'").replace('"', '\\"').replace("`", "\\`")

    command = f"""
getent passwd {name} > /dev/null && echo '{name} existiert schon, Abbrcuh' && exit 1
groupadd {name}
useradd -m -d /home/{name} -c {name} -g {name} -s /bin/bash {name} -G cdrom,plugdev,sambashare
echo {name}:{pwd} | chpasswd"""

    print(command, file=create)

def delete_user(name, delete):
    logger.info(f"Lösche: {name}")
    if verbose: print(f"echo lösche {name}", file=delete)
    print(f"userdel -r {replace_normalize(name).lower()}", file=delete)


def create_credentials():
    logger.info("Creating credentials sheet")
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet["A1"] = "Username"
    sheet["B1"] = "Password"
    return workbook, sheet


def add_credential(sheet, name, pwd, row):
    logger.info(f"Password for {name} added")
    sheet[f"A{row}"] = name
    sheet[f"B{row}"] = pwd

def generate_pwd(klasse, nummer, kv):
    logger.info("generiere Passwort")
    rand = ["!%&(),._-=^#"[random.randint(0,11)] for i in range(3)]
    return f"{klasse}{rand[0]}{nummer}{rand[1]}{kv}{rand[2]}"

def read_excel(file):
    try:
        b = openpyxl.load_workbook(file)
        s = b[b.sheetnames[0]]
        for row in s.iter_rows():
            if row[0].value is not None and row[0].value != "Klasse":
                yield [row[i].value for i in range(3)]
    except:
        logger.error("Datei nicht vorhanden")

def save_excel(wb):
    logger.info("pwd speichern")
    wb.save("user_credentials.xlsx")

def create_skripts(file):
    logger.info("Skriptgenerierung startet")

    cred_wb, cred_sh = create_credentials()

    with open("create_user.sh", "w") as create, open("delete_user.sh", "w") as delete:
        print("#!/bin/bash", file=create)
        print("set -e", file=create)
        print("#!/bin/bash", file=delete)
        print("set -e", file=delete)

        print("mkdir /home/klassen", file=create)

        create_delete_user("lehrer", create, delete)
        add_credential(cred_sh, "lehrer", "lehrer", 2)
        create_delete_user("seminar", create, delete)
        add_credential(cred_sh, "seminar", "seminar", 3)

        r = 4
        for i in read_excel(file):
            i[0] = f"k{i[0]}"
            pwd = generate_pwd(str(i[0]).lower(), str(i[1]).lower(), str(i[2]).lower())
            create_user(str(i[0]), create, pwd)
            delete_user(str(i[0]), delete)
            add_credential(cred_sh, str(i[0]), pwd, r)
            r += 1

    save_excel(cred_wb)
    logger.info("Alles erstellt")

def start_logging(verbose, quiet):
    if quiet:
        logger.setLevel(logging.ERROR)
    elif verbose:
        logger.setLevel(logging.DEBUG)
    else:
        logger.setLevel(logging.INFO)

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("filename", type=str, help="filename")
    parser.add_argument("-v", "--verbose", action="store_true", help="activates verbose mode")
    parser.add_argument("-q", "--quiet", action="store_true", help="activates quite mode")
    args = parser.parse_args()

    verbose = args.verbose
    quiet = args.quiet

    start_logging(verbose, quiet)
    try:
        create_skripts(args.filename)
    except FileNotFoundError:
        logger.error("File not found")



if __name__ == "__main__":
    main()